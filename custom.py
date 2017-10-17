import numpy as np
import pandas as pd
# from scipy.ndimage import filters
import os
import pdb


#%% Data management

def resample(data, ts, ts_ds, axis=0, empty_bin=0, method=np.mean, dtype=float):
    '''
    Downsamples data into new timestamps
    `data` is fitted into `ts_ds`. Throws out data points that do not fall
    within new sampling range.
    '''

    # only works for axis=0 at the moment

    # Make sure `data` is array
    if not isinstance(data, np.ndarray):
        data = np.array(data)
    
    # Setup
    dt_ds = len(ts_ds)
    data = data.swapaxes(-1, axis)  # Reorient `data` according to `axis`
    data_ds = empty_bin * np.ones(data.shape[:-1] + (dt_ds,), dtype=dtype)  # Initialize array for downsampled data
    
    # Index original timestamps to bin in new (downsampled) timestamps
    bin_ix = np.digitize(ts, ts_ds)
    
    # Downsample
    # Collect points in original dataset that fall within current 
    # downsampled bin. Average (or sum) these points.
    for bin_ds in np.arange(dt_ds):
        bin_pts = bin_ix == bin_ds + 1  # Indices within original dataset

        if np.any(bin_pts):
            data_ds[..., bin_ds] = method(data[..., bin_pts], axis=-1)
    
    return data_ds.swapaxes(-1, axis)


def interp_missing(incomplete_data):
    # Interpoloate values for NaN values in 1- or 2-d array x

    bad_ix = np.isnan(incomplete_data)
    good_ix = ~bad_ix
    all_ix = np.arange(incomplete_data.shape[1])
    
    # Interpolate data
    x = [all_ix[ix] for ix in bad_ix]
    xp = [all_ix[ix] for ix in good_ix]
    fp = [arr[ix] for arr, ix in zip(incomplete_data, good_ix)]
    missing_data = map(np.interp, x, xp, fp)
    
    # Create complete dataset
    interp_data = incomplete_data
    for d in np.arange(len(incomplete_data)):
        interp_data[d][x[d]] = missing_data[d]
    
    return interp_data


# interp_missing accomplishes waht xy_interp does
def xy_interp(x, y):
    # Interpolate values for NaN coordinates
    # Identify beginning and end of each gap in data. Interpolate 
    # data using linear interpolation between values around gaps.
    # 
    # Note: Fails if first or last element is nan.
    
    nan_ix = np.isnan(x)
    if np.any(nan_ix != np.isnan(y)):
        print 'NaN do not align between x and y coordinates'

    # Find indices where data is missing
    holes = np.diff(np.int_(nan_ix))
    hole0 = np.where(holes == 1)
    hole1 = np.where(holes == -1)

    # just need to first element in the returned tuple
    hole0 = hole0[0]
    hole1 = hole1[0]

    # Interpolate missing coordinates
    x_new = x.copy()
    y_new = y.copy()
    for i0, i1 in zip(hole0, hole1):
        x_new[i0+1:i1+1] = np.linspace(x[i0], x[i1+1], num=i1-i0+2)[1:-1]
        y_new[i0+1:i1+1] = np.linspace(y[i0], y[i1+1], num=i1-i0+2)[1:-1]
    
    return x_new, y_new


#%% Stats

def bootstrap(data, num_samples, statistic, alpha=0.1):
    n = len(data)
    idx = np.random.choice(data, n, (num_samples, n))
    samples = data[idx]
    stat = np.sort(statistic(samples, 1))
    return (stat[int((alpha/2.0)*num_samples)],
            stat[int((1-alpha/2.0)*num_samples)])


def zscore_base(data, baseline, axis=0):
    if baseline:
        if type(baseline) is int:
            b0 = 0
            b1 = baseline
        elif len(baseline) == 2 and\
             type(baseline[0]) is int:
            b0 = baseline[0]
            b1 = baseline[1]
        else:
            print "Incorrect baseline parameter"
            return

    data = np.swapaxes(data, axis, 0)
    base_avg = data[b0:b1, ...].mean(axis=0, keepdims=True)
    base_std = data[b0:b1, ...].std(axis=0, keepdims=True)
    data_z = (data - base_avg.repeat(data.shape[0], axis=0)) / base_std.repeat(data.shape[0], axis=0)

    return np.swapaxes(data_z, 0, axis)


def bh_correction(p_vals, fdr=0.05, asmask=False):
    '''
    Multiple comparisons correction by controlling false discovery rate.
    Assumes tests are independent of each other, eg, A vs B, C vs D, ...
    '''
    
    # Parameters
    # p_vals:  p_values obtained from multiple comparisons
    # fdr:    false discovery rate (FDR)
    # 
    # Returns
    # Array of indices corresponding to p_values that remain significant
    
    p_ord = np.sort(p_vals)
    p_ord_ix = np.argsort(p_vals)

    m = len(p_ord)
    threshold = fdr * (np.arange(m) + 1) / m
    p_small = np.where(p_ord < threshold)[0]
    if p_small.size:
        cutoff = p_small[-1]
        
        if asmask:
            significant_ix = np.zeros(p_vals.shape, dtype=bool)
            significant_ix[p_ord_ix[:cutoff + 1]] = True
        else:
            significant_ix = p_ord_ix[:cutoff + 1]

        return significant_ix
    else:
        return None


#%% Plotting

def style(key='default'):
    import matplotlib as mpl
    from cycler import cycler

    if key == 'default':
        pass
    elif key == 'succulent':
        mpl.rcParams['axes.prop_cycle'] = cycler(color=[
            '#FB6A64',
            '#654756',
            '#698184',
            '#B5C5C7',
            '#D8E1DB',
            '#FEC8C7',
        ])
    elif key == 'color-sea':
        # https://www.design-seeds.com/category/wander/sea/
        mpl.rcParams['axes.prop_cycle'] = cycler(color=[
            '#C6C1CB',
            '#283040',
            '#3C5465',
            '#638689',
            '#B5D5D8',
            '#E0E4E9',
        ])
    elif key == 'color-gaze':
        # https://www.design-seeds.com/category/wander/sea/
        mpl.rcParams['axes.prop_cycle'] = cycler(color=[
            '#41474B',
            '#1E1E23',
            '#25355F',
            '#4C6FA6',
            '#85BEDC',
            '#EDE1CF',
        ])
    elif key == 'summer-sky':
        # https://www.design-seeds.com/category/wander/sea/
        mpl.rcParams['axes.prop_cycle'] = cycler(color=[
            '#DCCFC1',
            '#233B4D',
            '#37617B',
            '#6C98AE',
            '#BCD1DD',
            '#E3F0F1',
        ])
    elif key == 'color-view':
        # https://www.design-seeds.com/category/wander/sea/
        mpl.rcParams['axes.prop_cycle'] = cycler(color=[
            '#84A19E',
            '#587674',
            '#B25949',
            '#E09A80',
            '#DEDAC5',
            '#E0E0DA',
        ])
    elif key == 'flora-tones':
        # https://www.design-seeds.com/category/wander/sea/
        mpl.rcParams['axes.prop_cycle'] = cycler(color=[
            '#DDBC90',
            '#06201E',
            '#304C43',
            '#C2C6C5',
            '#D6DBD9',
            '#E9E8E4',
        ])
    elif key == 'creature-color':
        # https://www.design-seeds.com/category/wander/sea/
        mpl.rcParams['axes.prop_cycle'] = cycler(color=[
            '#F0F0F0',
            '#E6E6EC',
            '#444444',
            '#D79763',
            '#FBCD8D',
            '#EFEFE4',
        ])
    elif key == 'color-creature':
        # https://www.design-seeds.com/category/wander/sea/
        mpl.rcParams['axes.prop_cycle'] = cycler(color=[
            '#EDEFF0',
            '#E1DFDF',
            '#D8C8AD',
            '#74786F',
            '#A19E50',
            '#D1D588',
        ])
    elif key == 'color-nature':
        # https://www.design-seeds.com/category/wander/sea/
        mpl.rcParams['axes.prop_cycle'] = cycler(color=[
            '#CBE8F3',
            '#213B3B',
            '#305045',
            '#6F987A',
            '#B1D4A6',
            '#E6E4CA',
        ])
    elif key == 'foraged-hues':
        # https://www.design-seeds.com/category/wander/sea/
        mpl.rcParams['axes.prop_cycle'] = cycler(color=[
            '#CECFDC',
            '#AEBED1',
            '#6A7D7A',
            '#4B4546',
            '#CF6C57',
            '#ECD6DC',
        ])
    elif key == 'barn-tones':
        # https://www.design-seeds.com/category/wander/sea/
        mpl.rcParams['axes.prop_cycle'] = cycler(color=[
            '#E8E8E8',
            '#2C2C2B',
            '#4C3332',
            '#763635',
            '#903A3B',
            '#AE5D5A',
        ])
    elif key == 'still-tones':
        # https://www.design-seeds.com/category/wander/sea/
        mpl.rcParams['axes.prop_cycle'] = cycler(color=[
            '#EAE8E4',
            '#D1CBC9',
            '#7F8A7F',
            '#474247',
            '#5A5979',
            '#8E8DAF',
        ])
    elif key == 'fresh-hues':
        # https://www.design-seeds.com/category/wander/sea/
        mpl.rcParams['axes.prop_cycle'] = cycler(color=[
            '#87B083',
            '#2E503D',
            '#DFAC5A',
            '#EBD175',
            '#F0EEA5',
            '#F1F9CF',
        ])
    elif key == 'market-hues':
        # https://www.design-seeds.com/category/wander/sea/
        mpl.rcParams['axes.prop_cycle'] = cycler(color=[
            '#BA2F27',
            '#441825',
            '#6A89B1',
            '#C5C4D6',
            '#DDD7E3',
            '#EAE3E4',
        ])
    elif key == 'coor-sip':
        # https://www.design-seeds.com/category/wander/sea/
        mpl.rcParams['axes.prop_cycle'] = cycler(color=[
            '#6EB6BE',
            '#387277',
            '#373737',
            '#664C3E',
            '#D98F5E',
            '#F0D8BB',
        ])
    elif key == 'color-serve':
        # https://www.design-seeds.com/category/wander/sea/
        mpl.rcParams['axes.prop_cycle'] = cycler(color=[
            '#C1C6CD',
            '#242F3E',
            '#513F41',
            '#CCB7A9',
            '#E2D5CE',
            '#F0EFE8',
        ])  
    elif key == 'color-collect':
        # https://www.design-seeds.com/category/wander/sea/
        mpl.rcParams['axes.prop_cycle'] = cycler(color=[
            '#FACFAA',
            '#DE8864',
            '#17311F',
            '#2B6D39',
            '#6BA085',
            '#AACDCA',
        ])  
    elif key == 'shelved-hues':
        # https://www.design-seeds.com/category/wander/sea/
        mpl.rcParams['axes.prop_cycle'] = cycler(color=[
            '#F5EEF0',
            '#E7D3D5',
            '#C49A63',
            '#67575A',
            '#6E7175',
            '#B2C1BA',
        ]) 
        
        
def nx_to_pydot(G, pydot_file=None, ext='raw', iplot=True, prog='neato'):
    
    import pydot
    
    if G.is_directed():
        graph_type = 'digraph'
    else:
        graph_type = 'graph'

    # Create pydot graph
    P = pydot.Dot(G.name, graph_type=graph_type)

    for n, nodedata in G.nodes(data=True):
        node = pydot.Node(str(n), **nodedata)
        P.add_node(node)

    for u, v, edgedata in G.edges(data=True):
        edge = pydot.Edge(str(u), str(v), **edgedata)
        P.add_edge(edge)

    # Save pydot
    if pydot_file:
        P.write(pydot_file, prog=prog, format=ext)

    return P


def hinton(matrix, max_weight=None, ax=None):
    """
    Draw Hinton diagram for visualizing a weight matrix.
    """
    
    import matplotlib.pyplot as plt

    ax = ax if ax is not None else plt.gca()

    if not max_weight:
        max_weight = 2 ** np.ceil(np.log(np.abs(matrix).max()) / np.log(2))

    ax.patch.set_facecolor('gray')
    ax.set_aspect('equal', 'box')
    ax.xaxis.set_major_locator(plt.NullLocator())
    ax.yaxis.set_major_locator(plt.NullLocator())

    for (x, y), w in np.ndenumerate(matrix):
        color = 'white' if w > 0 else 'black'
        size = np.sqrt(np.abs(w) / max_weight)
        rect = plt.Rectangle([x - size / 2, y - size / 2], size, size,
                             facecolor=color, edgecolor=color)
        ax.add_patch(rect)

    ax.autoscale_view()
    ax.invert_yaxis();

# def animate(movie, axis=0, cmap=None, clim=None, colorbar=False,
#             fps=30, plot=True, count=False, interpolation='none',
#             save=False, save_file='movie.mp4', title="Python movie", artist="Randall Ung"):
#     # Viability check
# #     if movie.ndim != 3:
# #         raise ValueError("Input array must be 3-dimensional.")
    
#     import matplotlib.pyplot as plt
#     from matplotlib import animation
    
#     if save or not plot:
#         istate = plt.isinteractive()
#         if istate: plt.ioff()
    
#     dt, dy, dx = movie.shape
    
#     fig, ax = plt.subplots(figsize=(6*dx/dy, 6))
#     im = ax.imshow(movie[0, ...], cmap=cmap, interpolation=interpolation)
    
#     if count: ax.set_title(0)
#     ax.set_ylim([movie.shape[1] - 1, 0])
#     ax.set_xlim([0, movie.shape[2] - 1])
#     ax.get_xaxis().set_visible(False)
#     ax.get_yaxis().set_visible(False)
#     ax.axis('image')
#     fig.tight_layout()

#     if clim: im.set_clim(clim)
#     if colorbar: plt.colorbar()
    
#     def init():
#         im.set_data(movie[0])
#         return im,
    
#     def ani(i):
#         if count: ax.set_title(str(i))
#         im.set_data(movie[i])
#         return im,
    
#     anim = animation.FuncAnimation(fig, ani, init_func=init, frames=dt, interval=1000/fps, blit=True)
    
#     if save:
#         metadata = dict(title=title,
#                         artist=artist)
#         FFwriter = animation.FFMpegWriter(fps=fps, metadata=metadata)
#         anim.save(save_file, writer=FFwriter)
            
#     if save or not plot:
#         if istate: plt.ion()
            
#     return anim

def animate(movies, axis=0, cmap=None, clim=None, colorbar=False,
            fps=30, plot=True, count=False, interpolation='none', width_ratios=None,
            save=False, save_file='movie.mp4', title="Python movie", artist="Randall Ung"):
    # Viability check
#     if movie.ndim != 3:
#         raise ValueError("Input array must be 3-dimensional.")
    
    import matplotlib.pyplot as plt
    from matplotlib import animation
    import matplotlib.gridspec as gridspec
    
    
    if save or not plot:
        istate = plt.isinteractive()
        if istate: plt.ioff()
    
    if not isinstance(movies, tuple):
        movies = (movies, )
    nMov = len(movies)
    dt = len(movies[0])
    if not all([len(movie) == dt for movie in movies]):
        raise ValueError("Not all movies match in frames")

    fig = plt.figure()
    gs = gridspec.GridSpec(1, len(movies))
    if width_ratios: gs.set_width_ratios(width_ratios)
    axs = [fig.add_subplot(g) for g in gs]
    ims = [ax.imshow(movie[0, ...], cmap=cmap, interpolation=interpolation) for movie, ax in zip(movies, axs)]
    for movie, ax in zip(movies, axs):
        # im = ax.imshow(movie[0, ...], cmap=cmap, interpolation=interpolation)
        # ims.append(im)
        ax.set_ylim([movie.shape[1] - 1, 0])
        ax.set_xlim([0, movie.shape[2] - 1])
        ax.get_xaxis().set_visible(False)
        ax.get_yaxis().set_visible(False)
        ax.axis('image')

    if count: fig.suptitle('0')
    fig.tight_layout()

    if clim:
        for im in ims: im.set_clim(clim)
    # if colorbar:
    #     fig.colorbar()
    
    def init():
        for movie, im in zip(movies, ims):
            im.set_data(movie[0])
        return ims
    
    def ani(i):
        if count: fig.suptitle(str(i))
        for movie, im in zip(movies, ims):
            im.set_data(movie[i])
        return ims
    
    anim = animation.FuncAnimation(fig, ani, init_func=init, frames=dt, interval=1000/fps, blit=True)
    
    if save:
        metadata = dict(title=title,
                        artist=artist)
        FFwriter = animation.FFMpegWriter(fps=fps, metadata=metadata)
        anim.save(save_file, writer=FFwriter)
            
    if save or not plot:
        if istate: plt.ion()
            
    return anim

def make_mov(filename, movie, roi=None, roi_color=[0, 1, 0], cmap=None, vmin=0, vmax=127, fps=30, dpi=100, title='', comments=''):
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.animation as manimation

    dt, dy, dx = movie.shape

    # Setup ROI (if applicable)
    multiroi = False

    if roi is not None:
        if isinstance(roi, str) and os.path.isfile(roi):
            # parameter is filename
            polyroi = np.loadtxt(roi, delimiter=',').astype('int32')
        else:
            polyroi = np.array(roi)
            if len(polyroi.shape) == 2:
                # parameter is single roi
                pass
            else:
                # parameter is series of roi
                if len(polyroi) != dt:
                    print("Length of movie and ROI do not match")
                    return
                else:
                    multiroi = True
    else:
        pass

    # Create figure
    fig, ax = plt.subplots()
    if roi:
        line, = ax.plot(polyroi[0][0], polyroi[0][1], color=roi_color)
    im = ax.imshow(np.zeros((dy, dx)), cmap=cmap, vmin=vmin, vmax=vmax)
    ax.get_xaxis().set_visible(False)
    ax.get_yaxis().set_visible(False)
    ax.axis('image')
    fig.subplots_adjust(left=0, bottom=0, right=1, top=1, wspace=0, hspace=0)

    # Movie writer
    # from https://matplotlib.org/examples/animation/moviewriter.html
    FFMpegWriter = manimation.writers['ffmpeg']
    metadata = dict(
        title=title,
        artist="Randall Ung",
        comment=comments)
    writer = FFMpegWriter(fps=fps, metadata=metadata)

    with writer.saving(fig, filename, dpi):
        for frame in xrange(len(movie)):
            im.set_data(movie[frame])
            if multiroi: line.set_data(roi[frame].T)
            writer.grab_frame(facecolor='k')

    print('Movie finished.')

def play(movie, movie2=None, frames=None, roi=None, roi_color=[0, 1, 0], text=None,
         axis=0, cmap=None, clim=None, colorbar=False,
         plot=True, interpolation='none'):
    '''
    roi: if multiple ROIs, first dimension should be 2 (i.e., X and Y values)

    TODO: play button
    TODO: play only certain frames defined by `frames` input
    '''

    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    import platform
    import cv2
    import Tkinter as tk


    def playback(fps):
        pass

    class Viewer(tk.Frame):

        def __init__(self, parent):
            ew = 6

            tk.Frame.__init__(self, parent)
            self.parent = parent
            parent.grid_columnconfigure(0, weight=1)
            parent.grid_rowconfigure(0, weight=1)
    
            dt, dy, dx = movie.shape
            
            # Create ROI mask if defined
            self.multiroi = False
            
            if roi is not None:
                if isinstance(roi, str) and os.path.isfile(roi):
                    # parameter is filename
                    polyroi = np.loadtxt(roi, delimiter=',').astype('int32')
                else:
                    polyroi = np.array(roi)
                    if len(polyroi.shape) == 2:
                        # parameter is single roi
                        pass
                    else:
                        # parameter is series of roi
                        if len(polyroi) != dt:
                            print("Length of movie and ROI do not match")
                            return
                        else:
                            self.multiroi = True
            else:
                pass

            # Viewing window
            fig, self.ax = plt.subplots(figsize=(6*dx/dy, 6))
            self.im = self.ax.imshow(movie[0], cmap=cmap, interpolation=interpolation)
            if roi is not None:
                if self.multiroi:
                    self.line, = self.ax.plot(polyroi[0][0], polyroi[0][1], color=roi_color)
                else:
                    self.ax.plot(polyroi[:, 0], polyroi[:, 1], color=roi_color)
            
            self.ax.set_ylim([movie.shape[1] - 1, 0])
            self.ax.set_xlim([0, movie.shape[2] - 1])
            self.ax.axis('off')
            fig.subplots_adjust(left=0, bottom=0, right=1, top=1, wspace=0, hspace=0)
            if clim: self.im.set_clim(clim)
            if colorbar: fig.colorbar(self.im, cax=self.ax)

            self.canvas = FigureCanvasTkAgg(fig, parent)
            self.canvas.show()
            self.canvas.draw()
            self.canvas.get_tk_widget().grid(row=0, column=0, sticky='wens')

            # Slider
            frame_tools = tk.Frame(parent)
            frame_tools.grid(row=1, column=0, sticky='we')
            # frame_tools.grid_columnconfigure(0, weight=1)
            # frame_tools.grid_columnconfigure(1, weight=1)
            frame_tools.grid_columnconfigure(2, weight=10)
            # frame_tools.grid_columnconfigure(3, weight=1)

            self.var_playing = tk.BooleanVar()

            self.entry_fps = tk.Entry(frame_tools, justify='center', width=ew)
            self.button_play = tk.Button(frame_tools, text=u"\u25B6", width=1, command=self.play_mov)
            self.scale = tk.Scale(frame_tools, orient='horizontal', from_=0, to=dt-1,
                showvalue=False, command=lambda val: self.update(val, 'scale'))
            if platform.system() == 'Linux':
                self.scale.bind_all('<Button-4>', self.mousewheel)
                self.scale.bind_all('<Button-5>', self.mousewheel)
            else:
                self.scale.bind_all('<MouseWheel>', self.mousewheel)
            self.entry_frame = tk.Entry(frame_tools, justify='center', width=ew)#, textvariable=self.var_frame)
            self.entry_frame.bind('<Return>', lambda ev: self.update(self.entry_frame.get(), 'entry'))

            self.entry_fps.grid(row=0, column=0, sticky='wens')
            self.button_play.grid(row=0, column=1, sticky='wens')
            self.scale.grid(row=0, column=2, sticky='wens')
            self.entry_frame.grid(row=0, column=3, sticky='wens')

            self.entry_fps.insert(0, 30)

        def mousewheel(self, event):
            # https://stackoverflow.com/questions/17355902/python-tkinter-binding-mousewheel-to-scrollbar
            df = 0
            if platform.system() == 'Linux':
                if event.num == 5:
                    df = 1
                if event.num == 4:
                    df = -1
            else:
                df = -event.delta / 120

            val = self.scale.get()
            self.scale.set(val + df)

        def play_mov(self):
            playing = self.var_playing.get()
            if playing:
                self.button_play['text'] = u'\u25B6'
                fps = int(self.entry_fps.get())

                # Threading
            else:
                self.button_play['text'] = u'\u23F8'

            self.var_playing.set(not playing)
        
        def update(self, frame, src):
            n = int(frame)
            if src != 'entry':
                self.entry_frame.delete(0, tk.END)
                self.entry_frame.insert(0, n)
            if src != 'scale':
                self.scale.set(n)            

            self.im.set_data(movie[n])
            if self.multiroi: self.line.set_data(roi[n].T)
            if text is not None: self.ax.set_title(text[n])
            self.canvas.draw_idle()

    root = tk.Tk()
    root.wm_title("Custom player")
    app = Viewer(root)
    root.mainloop()


#%% Random functions

def activity_map(x, y, sig, binsize=1, buffer=1, plot=False, sigma=0.5, truncate=4, cmap='jet'):
    # Creates activity heatmap
    # Inputs:
    #  x:  x-coordinate of path
    #  y:  y-coordinate of path
    #  sig: neural signal for each point along x,y-path
    #
    # Returns:
    #  heatmap: 2-dimensional array with entries corresponding to bins 
    #           in 2-d space. Values correspond to average neural 
    #           activity within bin.
    #  xy_bin:  Meshgrid corresponding to heatmap.
    #  [fig,
    #   ax,
    #   im]:    figure components
    #
    # Notes: x, y, sig must have same dimension. x, y must contain only 
    # numerical values. Use
    
    # Create bins
    # Bins are created so that bounds are multiples of 'binsize'
    x_bnds = binsize * np.arange(np.floor(x.min()/binsize)-buffer, np.ceil(x.max()/binsize)+1+buffer)
    y_bnds = binsize * np.arange(np.floor(y.min()/binsize)-buffer, np.ceil(y.max()/binsize)+1+buffer)

    x_bin_num = len(x_bnds) - 1
    y_bin_num = len(y_bnds) - 1
    heatmap = np.zeros((x_bin_num, y_bin_num))

    # Go through each bin to determine points within bin 
    # and average signal value for those points.
    for i, (x0, x1) in enumerate(zip(x_bnds[:-1], x_bnds[1:])):
        goodX = (x >= x0) & (x <  x1)               # index of points within bin's x-range
        for j, (y0, y1) in enumerate(zip(y_bnds[:-1], y_bnds[1:])):
            goodY = (y >= y0) & (y < y1)            # index of points within bin's y-range
            goodXY = goodX & goodY                      # index of points within both bin's x- & y-range
            if np.any(goodXY):
                # Bin value as average signal
                heatmap[i, j] = np.mean(sig[goodXY])
    
    # x,y-coordinates to plot
    X, Y = np.meshgrid(x_bnds, y_bnds)
    
    # Plot
    if plot:
        # Apply filter to smooth heatmap
        heatmap_filtered = filters.gaussian_filter(heatmap, sigma, truncate=truncate)
        data = heatmap_filtered.transpose()
        
        # Set transparent background
        C = np.ma.array(data, mask=data == 0)
        cmap = cm.get_cmap(cmap)
        cmap.set_bad('w', 0)
        
        fig, ax = plt.subplots()
        im = ax.pcolormesh(X, Y, C, cmap=cmap)
        fig.colorbar(im)
        ax.axis('equal')
        ax.axis('off')
        ax.set_xlim(X[0, 0], X[0, -1])
        ax.set_ylim(Y[0, 0], Y[-1, 0])
        
        plt.show()
    
        return heatmap, X, Y, [fig, ax, im]

    return heatmap, X, Y


# def etho_extract(filename):
#     from openpyxl import load_workbook
    
#     # Import Excel file
#     if not filename:
#         return -1
    
#     wb = load_workbook(filename)  # Workbook (file)
#     ws_names = wb.sheetnames      # Names of sheets
#     ws = wb[ws_names[0]]          # First worksheet

#     data_row = int(ws.rows[0][1].value) # first row with data
#     num_rows = ws.max_row
    
#     # Find number of columns
#     # Number identified can be greater than actual if iteration is used.
#     num_cols = 0
#     for cell in ws.rows[data_row-2]:
#         if cell.value is None:
#             break
#         else:
#             num_cols += 1
    
#     # Data variables
#     data_labels = [cell.value for cell in ws.rows[data_row-2]]
    
#     # Gather data
#     data_arr = np.empty((num_rows-data_row, num_cols), dtype=float)
#     for r, row in enumerate(ws.rows[data_row:]):
#         row_data = np.array([str(cell.value) for cell in row])
#         if row_data.dtype == 'float64':
#             data_arr[r, :] = row_data
#             print "no str"
#         else:
#             # Convert non numerical values to nan
#             data_arr[r, :] = np.genfromtxt(row_data)
            
#     return data_arr, data_labels

def etho_extract(filename, data_row=None, header=-2, index_col=1):
    """
    data_row:   Row that data starts at.
    header:     Location of header. Relative to data_row if negative.
    """

    if not os.path.isfile(filename):
        return -1
    
    ext = os.path.splitext(filename)[1]
    if ext in ['xls', 'xlsx']:
        # Import Excel file
        from openpyxl import load_workbook
        if data_row is None:
            wb = load_workbook(filename)  # Workbook (file)
            ws_names = wb.sheetnames      # Names of sheets
            ws = wb[ws_names[0]]          # First worksheet

            data_row = int(ws.rows[0][1].value) # first row with data
        
        skiprows = range(data_row)
        del skiprows[header]
        df = pd.read_excel(
            filename,
            skiprows=skiprows,
            na_values='-',
            index_col=index_col
        )
    elif ext in ['txt', 'csv']:
        raise IOError('Unable to process text files at the moment.')
    else:
        raise IOError('Unrecognized file type.')

    return df
