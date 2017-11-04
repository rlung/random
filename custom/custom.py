import numpy as np
import pandas as pd
import os
import pdb


#%% Data management

def resample(data, ts, ts_ds, axis=0, empty_bin=0, method=np.mean, dtype=float):
    '''
    Downsamples data into new timestamps
    `data` is fitted into `ts_ds`. Throws out data points that do not fall
    within new sampling range.
    '''

    # only works for axis=0 at the moment

    # Make sure `data` is array
    if not isinstance(data, np.ndarray):
        data = np.array(data)
    
    # Setup
    dt_ds = len(ts_ds)
    data = data.swapaxes(-1, axis)  # Reorient `data` according to `axis`
    data_ds = empty_bin * np.ones(data.shape[:-1] + (dt_ds,), dtype=dtype)  # Initialize array for downsampled data
    
    # Index original timestamps to bin in new (downsampled) timestamps
    bin_ix = np.digitize(ts, ts_ds)
    
    # Downsample
    # Collect points in original dataset that fall within current 
    # downsampled bin. Average (or sum) these points.
    for bin_ds in np.arange(dt_ds):
        bin_pts = bin_ix == bin_ds + 1  # Indices within original dataset

        if np.any(bin_pts):
            data_ds[..., bin_ds] = method(data[..., bin_pts], axis=-1)
    
    return data_ds.swapaxes(-1, axis)


def interp_missing(incomplete_data):
    # Interpoloate values for NaN values in 1- or 2-d array x

    bad_ix = np.isnan(incomplete_data)
    good_ix = ~bad_ix
    all_ix = np.arange(incomplete_data.shape[1])
    
    # Interpolate data
    x = [all_ix[ix] for ix in bad_ix]
    xp = [all_ix[ix] for ix in good_ix]
    fp = [arr[ix] for arr, ix in zip(incomplete_data, good_ix)]
    missing_data = map(np.interp, x, xp, fp)
    
    # Create complete dataset
    interp_data = incomplete_data
    for d in np.arange(len(incomplete_data)):
        interp_data[d][x[d]] = missing_data[d]
    
    return interp_data


# interp_missing accomplishes waht xy_interp does
def xy_interp(x, y):
    # Interpolate values for NaN coordinates
    # Identify beginning and end of each gap in data. Interpolate 
    # data using linear interpolation between values around gaps.
    # 
    # Note: Fails if first or last element is nan.
    
    nan_ix = np.isnan(x)
    if np.any(nan_ix != np.isnan(y)):
        print 'NaN do not align between x and y coordinates'

    # Find indices where data is missing
    holes = np.diff(np.int_(nan_ix))
    hole0 = np.where(holes == 1)
    hole1 = np.where(holes == -1)

    # just need to first element in the returned tuple
    hole0 = hole0[0]
    hole1 = hole1[0]

    # Interpolate missing coordinates
    x_new = x.copy()
    y_new = y.copy()
    for i0, i1 in zip(hole0, hole1):
        x_new[i0+1:i1+1] = np.linspace(x[i0], x[i1+1], num=i1-i0+2)[1:-1]
        y_new[i0+1:i1+1] = np.linspace(y[i0], y[i1+1], num=i1-i0+2)[1:-1]
    
    return x_new, y_new


#%% Stats

def bootstrap(data, num_samples, statistic, alpha=0.1):
    n = len(data)
    idx = np.random.choice(data, n, (num_samples, n))
    samples = data[idx]
    stat = np.sort(statistic(samples, 1))
    return (stat[int((alpha/2.0)*num_samples)],
            stat[int((1-alpha/2.0)*num_samples)])


def zscore_base(data, baseline, axis=0):
    if baseline:
        if type(baseline) is int:
            b0 = 0
            b1 = baseline
        elif len(baseline) == 2 and\
             type(baseline[0]) is int:
            b0 = baseline[0]
            b1 = baseline[1]
        else:
            print "Incorrect baseline parameter"
            return

    data = np.swapaxes(data, axis, 0)
    base_avg = data[b0:b1, ...].mean(axis=0, keepdims=True)
    base_std = data[b0:b1, ...].std(axis=0, keepdims=True)
    data_z = (data - base_avg.repeat(data.shape[0], axis=0)) / base_std.repeat(data.shape[0], axis=0)

    return np.swapaxes(data_z, 0, axis)


def bh_correction(p_vals, fdr=0.05, as_mask=False):
    '''Benjamini-Hochberg correction
    Multiple comparisons correction by controlling false discovery rate.
    Assumes tests are independent of each other, eg, A vs B, C vs D, ...
    '''
    
    # Parameters
    # p_vals:  p_values obtained from multiple comparisons
    # fdr:     false discovery rate (FDR)
    # 
    # Returns
    # Array of indices corresponding to p_values that remain significant
    
    p_ord = np.sort(p_vals)
    p_ord_ix = np.argsort(p_vals)

    m = len(p_ord)
    threshold = fdr * (np.arange(m) + 1) / m
    p_small = np.where(p_ord < threshold)[0]

    if p_small.size:
        # At least one p value made the cutoff
        cutoff = p_small[-1]
        
        if as_mask:
            significant_ix = np.zeros(p_vals.shape, dtype=bool)
            significant_ix[p_ord_ix[:cutoff + 1]] = True
        else:
            significant_ix = p_ord_ix[:cutoff + 1]

        return significant_ix
    else:
        # No p values made the cutoff
        return None


#%% Random functions


# def etho_extract(filename):
#     from openpyxl import load_workbook
    
#     # Import Excel file
#     if not filename:
#         return -1
    
#     wb = load_workbook(filename)  # Workbook (file)
#     ws_names = wb.sheetnames      # Names of sheets
#     ws = wb[ws_names[0]]          # First worksheet

#     data_row = int(ws.rows[0][1].value) # first row with data
#     num_rows = ws.max_row
    
#     # Find number of columns
#     # Number identified can be greater than actual if iteration is used.
#     num_cols = 0
#     for cell in ws.rows[data_row-2]:
#         if cell.value is None:
#             break
#         else:
#             num_cols += 1
    
#     # Data variables
#     data_labels = [cell.value for cell in ws.rows[data_row-2]]
    
#     # Gather data
#     data_arr = np.empty((num_rows-data_row, num_cols), dtype=float)
#     for r, row in enumerate(ws.rows[data_row:]):
#         row_data = np.array([str(cell.value) for cell in row])
#         if row_data.dtype == 'float64':
#             data_arr[r, :] = row_data
#             print "no str"
#         else:
#             # Convert non numerical values to nan
#             data_arr[r, :] = np.genfromtxt(row_data)
            
#     return data_arr, data_labels

def etho_extract(filename, data_row=None, header=-2, index_col=1):
    """
    data_row:   Row that data starts at.
    header:     Location of header. Relative to data_row if negative.
    """

    if not os.path.isfile(filename):
        return -1
    
    ext = os.path.splitext(filename)[1]
    if ext in ['xls', 'xlsx']:
        # Import Excel file
        from openpyxl import load_workbook
        if data_row is None:
            wb = load_workbook(filename)  # Workbook (file)
            ws_names = wb.sheetnames      # Names of sheets
            ws = wb[ws_names[0]]          # First worksheet

            data_row = int(ws.rows[0][1].value) # first row with data
        
        skiprows = range(data_row)
        del skiprows[header]
        df = pd.read_excel(
            filename,
            skiprows=skiprows,
            na_values='-',
            index_col=index_col
        )
    elif ext in ['txt', 'csv']:
        raise IOError('Unable to process text files at the moment.')
    else:
        raise IOError('Unrecognized file type.')

    return df
